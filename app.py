import streamlit as st
import pandas as pd
import io
from datetime import datetime
import numpy as np
from pathlib import Path
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font
warnings.filterwarnings('ignore')

class ExcelTransformer:
    def __init__(self, report_number=None):
        """
        Инициализация трансформера
        
        Args:
            report_number (str): Номер отчета для генерации кодов атрибутов
        """
        self.report_number = report_number or "R001"
        self.supported_extensions = ['.xlsx', '.xls', '.csv']
        self.report_types = ["Ручной", "Полуавтоматический", "Автоматический", "ИЛА"]
    
    def detect_data_type(self, values):
        """
        Автоматическое определение типа данных столбца
        
        Args:
            values: pandas Series с данными столбца
            
        Returns:
            str: тип данных ('текст', 'число', 'дата', 'флаг')
        """
        # Убираем пустые значения и NaN
        clean_values = values.dropna()
        if len(clean_values) == 0:
            return "текст"
        
        # Конвертируем в строки для анализа
        str_values = clean_values.astype(str).str.strip().str.lower()
        
        # Проверка на булевы значения (флаги)
        bool_indicators = {
            'да', 'нет', 'true', 'false', '1', '0', 'yes', 'no', 
            'y', 'n', 'вкл', 'выкл', 'on', 'off', 'активен', 'неактивен'
        }
        unique_values = set(str_values.unique())
        if unique_values.issubset(bool_indicators) and len(unique_values) <= 3:
            return "флаг"
        
        # Проверка на даты
        date_count = 0
        for val in clean_values:
            if self._is_date(val):
                date_count += 1
        
        if date_count / len(clean_values) > 0.7:  # 70% значений - даты
            return "дата"
        
        # Проверка на числа
        numeric_count = 0
        for val in clean_values:
            if self._is_numeric(val):
                numeric_count += 1
        
        if numeric_count / len(clean_values) > 0.8:  # 80% значений - числа
            return "число"
        
        return "текст"
    
    def _is_date(self, value):
        """Проверка, является ли значение датой"""
        if pd.isna(value):
            return False
            
        # Попробуем распарсить как дату
        date_formats = [
            '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y',
            '%d.%m.%y', '%d/%m/%y', '%y-%m-%d', '%d-%m-%y',
            '%Y.%m.%d', '%Y/%m/%d'
        ]
        
        str_val = str(value).strip()
        
        for fmt in date_formats:
            try:
                datetime.strptime(str_val, fmt)
                return True
            except ValueError:
                continue
                
        # Проверим pandas to_datetime
        try:
            pd.to_datetime(str_val, errors='raise')
            return True
        except:
            return False
    
    def _is_numeric(self, value):
        """Проверка, является ли значение числом"""
        if pd.isna(value):
            return False
            
        try:
            # Попробуем конвертировать в float
            float(str(value).replace(',', '.').replace(' ', ''))
            return True
        except ValueError:
            return False
    
    def load_from_uploaded_file(self, uploaded_file):
        """
        Загрузка данных из uploaded_file Streamlit
        
        Args:
            uploaded_file: файл, загруженный через st.file_uploader
            
        Returns:
            pandas.DataFrame: загруженные данные
        """
        try:
            file_extension = Path(uploaded_file.name).suffix.lower()
            
            if file_extension == '.csv':
                # Для CSV пробуем разные разделители и кодировки
                try:
                    df = pd.read_csv(uploaded_file, sep=',', encoding='utf-8')
                    if len(df.columns) > 1:
                        return df
                except:
                    pass
                
                try:
                    uploaded_file.seek(0)  # Сброс указателя файла
                    df = pd.read_csv(uploaded_file, sep=';', encoding='cp1251')
                    if len(df.columns) > 1:
                        return df
                except:
                    pass
                
                # Последняя попытка
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file)
                
            else:
                # Для Excel файлов
                df = pd.read_excel(uploaded_file)
            
            return df
            
        except Exception as e:
            raise Exception(f"Ошибка при загрузке файла: {str(e)}")
    
    def transform_to_metadata(self, df, report_type):
        """
        Преобразование DataFrame в метаданные атрибутов
        
        Args:
            df (pandas.DataFrame): исходные данные
            report_type (str): тип отчета (Ручной, Полуавтоматический, Автоматический, ИЛА)
            
        Returns:
            pandas.DataFrame: метаданные атрибутов
        """
        metadata_list = []
        
        for idx, column in enumerate(df.columns, 1):
            # Получаем данные столбца
            column_data = df[column]
            
            # Определяем тип данных
            data_type = self.detect_data_type(column_data)
            
            # Определяем значения по умолчанию в зависимости от типа отчета
            if report_type in ["Ручной", "Полуавтоматический"]:
                tech_algorithm_to_be = "Ручной ввод"
                data_source_type = "Ручное заполнение"
            else:  # Автоматический или ИЛА
                tech_algorithm_to_be = ""
                data_source_type = "База данных"
            
            # Связь с ИС для ИЛА
            system_connection = "ИЛА One" if report_type == "ИЛА" else ""
            
            metadata_record = {
                'ReportCode_info': '',  # Будет заполнено позже
                'Noreportfield_info': idx,
                'name': column,
                'description': '',
                'TechAsIs': '',
                'BussAlgorythm': '',
                'TechAlgorythm': tech_algorithm_to_be,
                'algorithms_change_info': 'нет',
                'dbobjectlink': '',
                'base_type_info': data_source_type,
                'related_it_system_info': system_connection,
                'reportfields_codes': '',
                'reportfields_names': '',
                'reportfields_parent_term': '',
                'reportfields_domain': '',
                'required_attribute_info': 'да',
                'base_type_report_field': data_type,
                'base_calc_ref_ind_info': 'Базовый',
                'codeTable_info': '',
                'example': '',
                'isToDelete_info': ''
            }
            
            metadata_list.append(metadata_record)
        
        metadata_df = pd.DataFrame(metadata_list)
        
        # Заполняем код атрибута после создания DataFrame
        metadata_df['ReportCode_info'] = metadata_df['Noreportfield_info'].apply(
            lambda x: f"{self.report_number}_{x}"
        )
        
        return metadata_df
    
    def create_excel_download(self, metadata_df):
        """
        Создание Excel файла для скачивания с заголовками и пользовательскими названиями
        
        Args:
            metadata_df (pandas.DataFrame): метаданные для сохранения
            
        Returns:
            bytes: данные Excel файла
        """
        output = io.BytesIO()
        
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Атрибут отчета"
        
        # Технические заголовки (скрытая строка)
        technical_headers = [
            'ReportCode_info', 'Noreportfield_info', 'name', 'description', 'TechAsIs', 
            'BussAlgorythm', 'TechAlgorythm', 'algorithms_change_info', 'dbobjectlink', 
            'base_type_info', 'related_it_system_info', 'reportfields_codes', 
            'reportfields_names', 'reportfields_parent_term', 'reportfields_domain', 
            'required_attribute_info', 'base_type_report_field', 'base_calc_ref_ind_info', 
            'codeTable_info', 'example', 'isToDelete_info'
        ]
        
        # Пользовательские заголовки (видимая строка)
        user_headers = [
            'Код атрибута отчета', 
            '№ атрибута отчета', 
            'Наименование атрибута', 
            'Описание атрибута',
            'Технический алгоритм AS IS', 
            'Бизнес-алгоритм AS IS',
            'Технический алгоритм TO BE', 
            'Алгоритм изменен', 
            'Физические атрибуты', 
            'Тип источника данных',
            'Связь с информационной системой', 
            'Код термина/терминов',
            'Наименование термина/терминов', 
            'Наименование родительской сущности термина/терминов', 
            'Домен термина/терминов', 
            'Обязательный атрибут для заполнения', 
            'Базовый тип атрибута (Текст, Число, Дата, Флаг)', 
            'Признак атрибута (Базовый, Расчетный, Справочный)', 
            'Наименование справочника', 
            'Примечание', 
            'Помечен к удалению'
        ]
        
        # Записываем технические заголовки в первую строку (скрытую)
        for col_idx, header in enumerate(technical_headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Скрываем первую строку
        ws.row_dimensions[1].hidden = True
        
        # Записываем пользовательские заголовки во вторую строку (видимую)
        for col_idx, header in enumerate(user_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            # Делаем заголовки полужирными
            cell.font = Font(bold=True)
        
        # Закрепляем первые две строки
        ws.freeze_panes = ws.cell(row=3, column=1)
        
        # Записываем данные начиная с третьей строки
        for row_idx, (_, row) in enumerate(metadata_df.iterrows(), 3):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Автоподбор ширины столбцов (учитываем все строки включая заголовки)
        for col_idx, column_letter in enumerate([chr(65 + i) for i in range(len(technical_headers))], 0):
            max_length = 0
            
            # Проверяем техническую строку
            if len(technical_headers) > col_idx:
                if len(str(technical_headers[col_idx])) > max_length:
                    max_length = len(str(technical_headers[col_idx]))
            
            # Проверяем пользовательскую строку
            if len(user_headers) > col_idx:
                if len(str(user_headers[col_idx])) > max_length:
                    max_length = len(str(user_headers[col_idx]))
            
            # Проверяем данные
            for row_idx in range(3, len(metadata_df) + 3):
                cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        wb.save(output)
        output.seek(0)
        return output.getvalue()

def main():
    # Настройка страницы без боковой панели
    st.set_page_config(
        page_title="Сгенерировать атрибутный состав",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Скрытие боковой панели
    st.markdown("""
    <style>
        .css-1d391kg {display: none}
        .css-1rs6os {display: none}
        .css-17eq0hr {display: none}
        [data-testid="stSidebar"] {display: none}
    </style>
    """, unsafe_allow_html=True)
    
    # Заголовок приложения
    st.title("📊 Сгенерировать атрибутный состав")
    st.markdown("### Преобразование горизонтальной структуры Excel в вертикальные метаданные атрибутов")
    st.markdown("---")
    
    # Инициализация session state
    if 'file_processed' not in st.session_state:
        st.session_state.file_processed = False
    if 'transformer' not in st.session_state:
        st.session_state.transformer = None
    if 'df' not in st.session_state:
        st.session_state.df = None
    if 'uploaded_file_name' not in st.session_state:
        st.session_state.uploaded_file_name = ""
    
    # Основной интерфейс
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.header("📁 Загрузка файла")
        uploaded_file = st.file_uploader(
            "Выберите Excel или CSV файл",
            type=['xlsx', 'xls', 'csv'],
            help="Поддерживаются форматы: Excel (.xlsx, .xls) и CSV (.csv)"
        )
    
    with col2:
        st.header("⚙️ Параметры")
        report_number = st.text_input(
            "Номер отчета",
            value="R001",
            help="Этот номер будет использован для генерации кодов атрибутов и названия файла"
        )
        
        report_type = st.selectbox(
            "Тип отчета",
            options=["Ручной", "Полуавтоматический", "Автоматический", "ИЛА"],
            help="Тип отчета влияет на заполнение технических полей метаданных"
        )
    
    if uploaded_file is not None:
        st.session_state.uploaded_file_name = uploaded_file.name
        
        # Создаем трансформер
        transformer = ExcelTransformer(report_number=report_number)
        
        try:
            # Загружаем данные
            df = transformer.load_from_uploaded_file(uploaded_file)
            
            # Сохраняем в session state
            st.session_state.transformer = transformer
            st.session_state.df = df
            st.session_state.file_processed = True
            
            st.success("✅ Файл успешно загружен!")
            
            # Показываем базовую информацию о файле
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Количество строк", len(df))
            with col2:
                st.metric("Количество столбцов", len(df.columns))
            with col3:
                st.metric("Номер отчета", report_number)
            
            # Кнопка для выгрузки атрибутного состава
            st.markdown("---")
            if st.button("🔄 Выгрузить атрибутный состав", type="primary", use_container_width=True):
                with st.spinner("Преобразование данных..."):
                    # Сброс указателя файла и преобразование
                    uploaded_file.seek(0)
                    df_fresh = transformer.load_from_uploaded_file(uploaded_file)
                    metadata_df = transformer.transform_to_metadata(df_fresh, report_type)
                    
                    # Создаем Excel файл для скачивания
                    excel_data = transformer.create_excel_download(metadata_df)
                    
                    # Генерируем имя файла
                    filename = f"{report_number}_атрибуты.xlsx"
                    
                    st.success("✅ Атрибутный состав успешно создан!")
                    
                    # Показываем статистику
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Создано атрибутов", len(metadata_df))
                    with col2:
                        type_stats = metadata_df['base_type_report_field'].value_counts()
                        most_common_type = type_stats.index[0] if len(type_stats) > 0 else "N/A"
                        st.metric("Основной тип", most_common_type)
                    with col3:
                        st.metric("Тип отчета", report_type)
                    
                    # Кнопка скачивания
                    st.download_button(
                        label="📥 Скачать атрибутный состав",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
                    
                    st.info("💡 **Совет**: Файл содержит скрытую строку с техническими заголовками для системной обработки и видимую строку с понятными названиями.")
        
        except Exception as e:
            st.error(f"❌ Ошибка при обработке файла: {str(e)}")
    
    # Подвал
    st.markdown("---")
    st.markdown("*Создано с помощью Streamlit* 🚀")

if __name__ == "__main__":
    main()
